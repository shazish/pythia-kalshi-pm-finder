"""
excel_reporter.py — Export classified Kalshi candidates to a professional Excel workbook.

Requires openpyxl:  sudo apt install python3-openpyxl
Falls back to CSV when openpyxl is not available.

Sheets produced:
  1. Opportunities  — candidates above the edge threshold (to_notify)
  2. All Results    — every classified market with routing decision
"""
import csv
import os
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Colour palette ──────────────────────────────────────────────────────────
_C = {
    "header_bg":   "1F3864",   # dark navy
    "header_fg":   "FFFFFF",
    "certain":     "C6EFCE",   # green fill
    "certain_fg":  "276221",
    "likely":      "FFEB9C",   # yellow fill
    "likely_fg":   "9C6500",
    "unclear":     "FCE4D6",   # orange fill
    "unclear_fg":  "843C0C",
    "anomaly":     "F4CCCC",   # red-tinted
    "near_miss":   "D9D9D9",   # grey — CERTAIN but failed validation or below threshold
    "near_miss_fg":"595959",
    "notify":      "D9EAD3",   # light green row tint for opportunities
    "subheader":   "D6E4F7",   # light blue for sub-labels
    "border":      "BFBFBF",
}

_THIN = None  # filled lazily after openpyxl import check


def _thin_border():
    global _THIN
    if _THIN is None:
        side = Side(style="thin", color=_C["border"])
        _THIN = Border(left=side, right=side, top=side, bottom=side)
    return _THIN


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _header_style():
    return {
        "font": _font(bold=True, color=_C["header_fg"], size=10),
        "fill": _fill(_C["header_bg"]),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _thin_border(),
    }


def _apply(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


# ── URL helpers ──────────────────────────────────────────────────────────────

def _market_url(r):
    """Return the web URL for the market on Kalshi or Polymarket."""
    from pipeline_logger import get_logger
    log = get_logger("excel_reporter")

    candidate = r.get("candidate", r)
    platform = candidate.get("platform", "Kalshi")
    if platform == "Polymarket":
        url = candidate.get("settlement_source_url", "")
        if not url:
            slug = candidate.get("slug", "")
            url = f"https://polymarket.com/event/{slug}" if slug else ""
        return url
    else:
        series_ticker = candidate.get("series_ticker", "") or candidate.get("event_ticker", "")
        if not series_ticker:
            log.warning("No series_ticker or event_ticker for ticker=%s — URL will be missing",
                        candidate.get("ticker", "?"))
            return ""
        return f"https://kalshi.com/markets/{series_ticker.lower()}"


def _ticker_cell(r):
    """Return a HYPERLINK formula so the Ticker cell links to the market page."""
    ticker = r["candidate"].get("ticker", "")
    url = _market_url(r)
    if url:
        safe_url = url.replace('"', "%22")
        safe_ticker = ticker.replace('"', "'")
        return f'=HYPERLINK("{safe_url}","{safe_ticker}")'
    return ticker


# ── Column definitions ───────────────────────────────────────────────────────

OPPORTUNITY_COLS = [
    ("Ticker",                 20, _ticker_cell),
    ("Title",                  45, lambda r: r["candidate"].get("title", "")),
    ("Category",               14, lambda r: r["candidate"].get("category", "")),
    ("Side",                    6, lambda r: r["classification"].get("high_confidence_side", "")),
    ("Market Price (c)",       14, lambda r: int(r["candidate"].get("implied_probability", 0) or 0)),
    ("Confidence %",           13, lambda r: r["classification"].get("confidence_score", "")),
    ("Classification",         14, lambda r: r["classification"].get("classification", "")),
    ("Status",                 24, lambda r: r.get("_opportunity_status", r.get("routing", ""))),
    ("Validation Errors",      40, lambda r: " | ".join(r["classification"].get("_validation_errors", []))),
    ("Edge %",                 10, lambda r: _pct(r.get("edge_after_fees"))),
    ("Ann. Edge %",            10, lambda r: _pct(r.get("annualized_edge"))),
    ("Fee Rate",               10, lambda r: _pct(r.get("fee_rate_used"))),
    ("Days to Close",          13, lambda r: r.get("days_to_close", "")),
    ("Urgency Score",          13, lambda r: r["candidate"].get("urgency_score", "")),
    ("Close Date",             12, lambda r: _date(r["candidate"].get("close_date", ""))),
    ("Suggested Size ($)",     16, lambda r: r.get("position_size_usd", "")),
    ("Volume",                 12, lambda r: int(r["candidate"].get("volume", 0) or 0)),
    ("Volume Anomaly",         22, lambda r: _anomaly_str(r["candidate"].get("volume_anomaly"))),
    ("Contradicting Signals",  35, lambda r: _contra_str(r["classification"].get("contradicting_signals", []))),
    ("Settlement Risk",        35, lambda r: r["classification"].get("settlement_risk", "")),
    ("What Would Change This", 40, lambda r: r["classification"].get("what_would_change_this", "")),
    ("Recent Developments",    40, lambda r: r["classification"].get("recent_developments", "")),
    ("Routing",                18, lambda r: r.get("routing", "")),
    ("Scan Type",              14, lambda r: r["candidate"].get("scan_type", "")),
    ("Scanned At",             20, lambda r: _date(r["candidate"].get("scanned_at", ""))),
]

ALL_RESULTS_COLS = OPPORTUNITY_COLS  # same columns, more rows


# ── Helpers ──────────────────────────────────────────────────────────────────

def _pct(val):
    if val is None:
        return ""
    return round(float(val) * 100, 2)


def _date(s):
    if not s:
        return ""
    return s[:10]  # YYYY-MM-DD


def _anomaly_str(anomaly):
    if not anomaly:
        return ""
    return (
        f"{anomaly['opposite_side']} @{anomaly['opposite_price']}c | "
        f"~${anomaly['implied_longshot_dollars']:,} implied "
        f"({anomaly['total_volume']:,} contracts)"
    )


def _contra_str(signals):
    if not signals:
        return ""
    return " | ".join(s.get("fact", "") for s in signals if s.get("fact"))


def _row_fill(cls):
    fills = {"CERTAIN": _C["certain"], "LIKELY": _C["likely"], "UNCLEAR": _C["unclear"]}
    color = fills.get(cls, "FFFFFF")
    return _fill(color)


def _row_font(cls):
    fgs = {"CERTAIN": _C["certain_fg"], "LIKELY": _C["likely_fg"], "UNCLEAR": _C["unclear_fg"]}
    return _font(color=fgs.get(cls, "000000"))


# ── Excel writer ─────────────────────────────────────────────────────────────

def _write_sheet(ws, col_defs, rows, title):
    """Write a single sheet: header + data rows."""
    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_defs))
    title_cell = ws.cell(row=1, column=1, value=title)
    _apply(title_cell,
           font=_font(bold=True, color=_C["header_fg"], size=12),
           fill=_fill(_C["header_bg"]),
           alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 22

    # Column headers (row 2)
    headers = [c[0] for c in col_defs]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        hs = _header_style()
        _apply(cell, **hs)
    ws.row_dimensions[2].height = 30

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=3, column=1)

    # Data rows
    for row_idx, record in enumerate(rows, start=3):
        cls = record.get("classification", {}).get("classification", "")
        has_anomaly = bool(record.get("candidate", {}).get("volume_anomaly"))
        is_near_miss = record.get("_is_near_miss", False)
        if is_near_miss:
            row_fill = _fill(_C["near_miss"])
            row_font = _font(color=_C["near_miss_fg"])
        elif has_anomaly:
            row_fill = _fill(_C["anomaly"])
            row_font = _row_font(cls)
        else:
            row_fill = _row_fill(cls)
            row_font = _row_font(cls)

        for col_idx, (_, _, extractor) in enumerate(col_defs, start=1):
            try:
                value = extractor(record)
            except Exception:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = row_font
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[row_idx].height = 45

    # Column widths
    for col_idx, (_, width, _) in enumerate(col_defs, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


def _write_legend(wb):
    """Add a Legend sheet explaining colour codes."""
    ws = wb.create_sheet("Legend")
    rows = [
        ("Colour", "Meaning"),
        ("Green (CERTAIN)", "Classifier assessed outcome as near-certain (≥95% confidence, ≥3 confirming signals, no contradicting signals). Actionable."),
        ("Grey (NEAR MISS)", "Classified CERTAIN but filtered — validation errors, edge below 3%, or already notified within 7 days. Review the Validation Errors column."),
        ("Yellow (LIKELY)", "Classifier assessed outcome as probable but not certain"),
        ("Orange (UNCLEAR)", "Classifier could not determine outcome with sufficient confidence"),
        ("Red tint", "Market has a VOLUME ANOMALY — large bets against the high-confidence side. Requires investigation."),
        ("", ""),
        ("Column", "Description"),
        ("Status", "OPPORTUNITY = actionable; NEAR MISS = CERTAIN but filtered (see Validation Errors)"),
        ("Validation Errors", "Why a CERTAIN candidate failed structural checks (e.g. fewer than 3 searches logged). Fix by re-running classification with stricter search logging."),
        ("Edge %", "Expected return after Kalshi fees as % of capital deployed"),
        ("Ann. Edge %", "Edge annualised by days to close (365 / days)"),
        ("Suggested Size ($)", "Kelly-criterion position size, capped at 5% of bankroll"),
        ("Volume Anomaly", "Opposite-side implied $ = volume × opposite_price / 100"),
        ("Contradicting Signals", "Facts found by classifier that argue against the high-confidence outcome"),
    ]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    for r_idx, (col_a, col_b) in enumerate(rows, start=1):
        ca = ws.cell(row=r_idx, column=1, value=col_a)
        cb = ws.cell(row=r_idx, column=2, value=col_b)
        if r_idx == 1 or col_a in ("Column",):
            for c in (ca, cb):
                c.font = _font(bold=True, color=_C["header_fg"])
                c.fill = _fill(_C["header_bg"])
        ca.alignment = Alignment(vertical="top")
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r_idx].height = 18


def _extract_near_misses(to_log):
    """
    Pull CERTAIN candidates that didn't notify (failed validation or below edge threshold)
    out of to_log so they can be shown on the Opportunities sheet as near-misses.
    """
    near_misses = []
    for r in to_log:
        cls = r.get("classification", {})
        if cls.get("classification") != "CERTAIN":
            continue
        routing = r.get("routing", "")
        if routing in ("skipped_validation_failed", "logged_below_threshold", "skipped_already_notified"):
            status_labels = {
                "skipped_validation_failed": "NEAR MISS — validation failed",
                "logged_below_threshold":    "NEAR MISS — edge below threshold",
                "skipped_already_notified":  "NEAR MISS — already notified",
            }
            copy = dict(r)
            copy["_is_near_miss"] = True
            copy["_opportunity_status"] = status_labels.get(routing, routing)
            near_misses.append(copy)
    return near_misses


def export_excel(to_notify, to_log, output_path, mode_label=""):
    """
    Write a two-sheet Excel workbook.

    Opportunities sheet: actionable rows (green) + CERTAIN near-misses (grey).
    Near-misses are CERTAIN candidates that were filtered by validation errors,
    edge threshold, or deduplication — shown so the sheet is never misleadingly empty.

    Args:
        to_notify:    list of opportunity dicts above edge threshold
        to_log:       list of all other classified results
        output_path:  .xlsx file path
        mode_label:   scan mode string appended to sheet titles (e.g. "deep").
                      Defaults to "finalize" when not provided.
    """
    from pipeline_logger import get_logger
    log = get_logger("excel_reporter")

    if not mode_label:
        mode_label = "finalize"

    log.info("export_excel() — notifying=%d, logging=%d, path=%s",
             len(to_notify), len(to_log), output_path)
    # Tag notify rows with their status
    for r in to_notify:
        r["_opportunity_status"] = "OPPORTUNITY"
        r["_is_near_miss"] = False

    near_misses = _extract_near_misses(to_log)
    opp_rows = to_notify + near_misses  # notified first, near-misses below

    if not OPENPYXL_AVAILABLE:
        _export_csv_fallback(to_notify, to_log, output_path)
        return output_path.replace(".xlsx", ".csv")

    wb = openpyxl.Workbook()

    # Sheet 1: Opportunities (actionable + near-misses)
    ws_opp = wb.active
    ws_opp.title = "Opportunities"
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    opp_title = (
        f"Kalshi {mode_label.capitalize()} Opportunities — {timestamp}  "
        f"({len(to_notify)} actionable  |  {len(near_misses)} near-miss)"
    )
    _write_sheet(ws_opp, OPPORTUNITY_COLS, opp_rows, opp_title)

    # Sheet 2: All Results
    ws_all = wb.create_sheet("All Results")
    all_rows = to_notify + to_log
    _write_sheet(
        ws_all,
        ALL_RESULTS_COLS,
        all_rows,
        f"All Classified Markets — {mode_label.capitalize()} — {timestamp}  ({len(all_rows)} total)",
    )

    # Sheet 3: Legend
    _write_legend(wb)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ── CSV fallback ─────────────────────────────────────────────────────────────

def _export_csv_fallback(to_notify, to_log, xlsx_path):
    """Write a plain CSV when openpyxl is not installed."""
    csv_path = xlsx_path.replace(".xlsx", ".csv")
    all_rows = to_notify + to_log
    # Use plain ticker for CSV (no formulas), append Market URL as extra column
    csv_cols = [
        ("Ticker", lambda r: r["candidate"].get("ticker", "")),
        *[(c[0], c[2]) for c in OPPORTUNITY_COLS[1:]],
        ("Market URL", _market_url),
    ]
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([c[0] for c in csv_cols])
        for record in all_rows:
            row = []
            for _, extractor in csv_cols:
                try:
                    row.append(extractor(record))
                except Exception:
                    row.append("")
            writer.writerow(row)
    print(f"[excel_reporter] openpyxl not available — wrote CSV fallback: {csv_path}")
    print("[excel_reporter] Install: sudo apt install python3-openpyxl")
    return csv_path


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json, sys
    classified_file = os.path.expanduser("~/.hermes/kalshi-tracker/cache/classified.json")
    if not os.path.exists(classified_file):
        print("No classified.json found.")
        sys.exit(1)

    with open(classified_file) as f:
        classified = json.load(f)

    from opportunity_manager import OpportunityManager
    mgr = OpportunityManager()
    to_notify, to_log = mgr.process(classified)

    out = os.path.expanduser(
        f"~/.hermes/kalshi-tracker/logs/kalshi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    result = export_excel(to_notify, to_log, out)
    print(f"Exported: {result}")
    print(f"  Opportunities: {len(to_notify)}")
    print(f"  All results:   {len(to_notify) + len(to_log)}")
